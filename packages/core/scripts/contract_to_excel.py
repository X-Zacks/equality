import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 读取 JSON 文件
with open(r'C:\Users\zz\Desktop\OCR Testing\ocr_results_incremental.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_header_style(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

def set_cell_style(cell):
    cell.alignment = left_align
    cell.border = thin_border

# ========== Sheet 1: 合同汇总 ==========
ws_summary = wb.active
ws_summary.title = '合同汇总'

# 标题行
ws_summary['A1'] = '合同信息汇总表'
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary.merge_cells('A1:H1')

# 表头
headers = ['序号', '合同编号', '合同名称', '甲方', '乙方', '生效日期', '到期日期', '合同金额', '货币', '状态']
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row=3, column=col, value=header)
    set_header_style(cell)

# 合同数据
contracts = [
    {
        'id': 1,
        'no': 'CW2679070-1',
        'name': 'Non-Technical Services Agreement (SOW)',
        'party_a': 'Lenovo HK Services Ltd',
        'party_b': 'eSOON China Limited',
        'start': '2024-12-31',
        'end': '2025-12-31',
        'amount': 833617.13,
        'currency': 'USD',
        'status': '有效'
    },
    {
        'id': 2,
        'no': 'CW2568992',
        'name': 'Splunk Enterprise 采购订单',
        'party_a': '联想（北京）有限公司',
        'party_b': '北京信诺时代科技发展有限公司',
        'start': '2023-05-01',
        'end': '2026-04-30',
        'amount': 5675604,
        'currency': 'CNY',
        'status': '有效'
    }
]

for row, contract in enumerate(contracts, 4):
    ws_summary.cell(row=row, column=1, value=contract['id'])
    ws_summary.cell(row=row, column=2, value=contract['no'])
    ws_summary.cell(row=row, column=3, value=contract['name'])
    ws_summary.cell(row=row, column=4, value=contract['party_a'])
    ws_summary.cell(row=row, column=5, value=contract['party_b'])
    ws_summary.cell(row=row, column=6, value=contract['start'])
    ws_summary.cell(row=row, column=7, value=contract['end'])
    ws_summary.cell(row=row, column=8, value=contract['amount'])
    ws_summary.cell(row=row, column=9, value=contract['currency'])
    ws_summary.cell(row=row, column=10, value=contract['status'])
    
    for col in range(1, 11):
        set_cell_style(ws_summary.cell(row=row, column=col))

# 设置列宽
col_widths = [8, 18, 35, 28, 28, 15, 15, 15, 10, 10]
for i, width in enumerate(col_widths, 1):
    ws_summary.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 2: JD SOW 详细信息 ==========
ws_jd = wb.create_sheet('JD-SOW-明细')

ws_jd['A1'] = 'JD-SOW Genesys 订阅服务明细'
ws_jd['A1'].font = Font(bold=True, size=14)
ws_jd.merge_cells('A1:G1')

# 基本信息
ws_jd['A3'] = 'SOW编号'
ws_jd['B3'] = 'CW2679070-1'
ws_jd['A4'] = '基础协议编号'
ws_jd['B4'] = 'CW2219441'
ws_jd['A5'] = '甲方'
ws_jd['B5'] = 'Lenovo HK Services Ltd'
ws_jd['A6'] = '乙方'
ws_jd['B6'] = 'eSOON China Limited'
ws_jd['A7'] = '服务内容'
ws_jd['B7'] = 'JD-Voice CC平台和LI Sales CC平台 1年期Genesys订阅服务'

for row in range(3, 8):
    ws_jd.cell(row=row, column=1).font = Font(bold=True)
    ws_jd.cell(row=row, column=1).fill = subheader_fill

# BOM 表
ws_jd['A9'] = 'JD-CC Voice Platform - Genesys Subscription BOM'
ws_jd['A9'].font = Font(bold=True)
ws_jd.merge_cells('A9:G9')

jd_bom_headers = ['序号', 'Building Block', '产品', '单位', '数量']
for col, header in enumerate(jd_bom_headers, 1):
    cell = ws_jd.cell(row=10, column=col, value=header)
    set_header_style(cell)

jd_bom_data = [
    (1, 'Base Package', 'Genesys Engage On Premise Base Package NR', 'Each', 2903),
    (2, 'Gplus Adapters (SAP, Siebel, etc)', 'v8.0 - Gplus Adapter SAP CRM - SUB', 'User', 100),
    (3, 'Framework & Intelligent Routing', 'v8.1- SNMP - SUB', 'Site', 1),
    (4, 'GVP', 'v8.5 - HA - Genesys Voice Platform - SUB', 'Port', 310),
    (5, 'GVP', 'v8.5 - Genesys Voice Platform (incl add\'l capability) - SUB', 'Port', 310),
    (6, 'GVP', 'v8.5 - AIModule for TTS - SUB', 'Port', 78),
    (7, 'GVP', 'v8.5 - HA - AIModule for TTS - SUB', 'Port', 78),
    (8, 'Web Call-back', 'v8.5 - Genesys Callback - SUB', 'User', 500),
    (9, 'WFM + Scheduling', 'v8.5 - Genesys Workforce Management - SUB', 'User', 821),
    (10, 'Gplus Adapters (SAP, Siebel, etc)', 'v8.5 - Gplus Adapter for Web Services - SUB', 'User', 200),
]

for row, item in enumerate(jd_bom_data, 11):
    for col, val in enumerate(item, 1):
        cell = ws_jd.cell(row=row, column=col, value=val)
        set_cell_style(cell)

# LI Sales BOM
ws_jd['A22'] = 'LI Sales CC Platform (JL) - Genesys Subscription BOM'
ws_jd['A22'].font = Font(bold=True)
ws_jd.merge_cells('A22:G22')

for col, header in enumerate(jd_bom_headers, 1):
    cell = ws_jd.cell(row=23, column=col, value=header)
    set_header_style(cell)

li_bom_data = [
    (1, 'Inbound Voice', 'Genesys Engage On Premise Base Package NR', 'Each', 852),
    (2, 'Web Call-back', 'v8.5 - Genesys Callback - SUB', 'User', 30),
    (3, 'GVP', 'v8.5 - Genesys Voice Platform (incl add\'l capability) - SUB', 'Port', 95),
    (4, 'GVP', 'v8.5 - HA - Genesys Voice Platform - SUB', 'Port', 95),
    (5, 'WFM + Scheduling', 'v8.5 - Genesys Workforce Management - SUB', 'User', 220),
    (6, 'Gplus Adapters (SAP, Siebel, etc)', 'v8.5 - Gplus Adapter for Web Services - SUB', 'User', 346),
]

for row, item in enumerate(li_bom_data, 24):
    for col, val in enumerate(item, 1):
        cell = ws_jd.cell(row=row, column=col, value=val)
        set_cell_style(cell)

# 报价汇总
ws_jd['A31'] = '续约报价汇总'
ws_jd['A31'].font = Font(bold=True)

quote_headers = ['序号', '项目名称', '数量', '续约周期', '2025年续约报价(USD)']
for col, header in enumerate(quote_headers, 1):
    cell = ws_jd.cell(row=32, column=col, value=header)
    set_header_style(cell)

quote_data = [
    (1, 'JD Production System Genesys Subscription Bom', '2903坐席', '2025.1.1-2025.12.31', ''),
    (2, 'JL Production System Genesys Subscription Bom', '852坐席', '2025.1.1-2025.12.31', 811845.46),
    (3, 'APJ Production System Genesys Subscription Bom', '476坐席', '2025.1.1-2025.12.31', ''),
    (4, '2022年Sub扩容75坐席renew', '75坐席', '2024.11.1-2025.10.31', 21771.67),
    ('', '续约报价总计', '', '', 833617.13),
]

for row, item in enumerate(quote_data, 33):
    for col, val in enumerate(item, 1):
        cell = ws_jd.cell(row=row, column=col, value=val)
        set_cell_style(cell)
        if row == 37:  # 总计行
            cell.font = Font(bold=True)

# 设置列宽
for i, width in enumerate([8, 45, 50, 15, 20], 1):
    ws_jd.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 3: Splunk 采购订单 ==========
ws_splunk = wb.create_sheet('Splunk-采购订单')

ws_splunk['A1'] = 'Splunk Enterprise 采购订单'
ws_splunk['A1'].font = Font(bold=True, size=14)
ws_splunk.merge_cells('A1:H1')

ws_splunk['A3'] = '订单编号'
ws_splunk['B3'] = 'CW2568992'
ws_splunk['A4'] = '生效日'
ws_splunk['B4'] = '签约日'
ws_splunk['A5'] = '联想合同编号'
ws_splunk['B5'] = 'CW2564504'
ws_splunk['A6'] = '供应商合同编号'
ws_splunk['B6'] = 'SINO-KJ-2023-72'

ws_splunk['A8'] = '甲方'
ws_splunk['B8'] = '联想（北京）有限公司'
ws_splunk['A9'] = '乙方'
ws_splunk['B9'] = '北京信诺时代科技发展有限公司'

for row in range(3, 10):
    ws_splunk.cell(row=row, column=1).font = Font(bold=True)
    ws_splunk.cell(row=row, column=1).fill = subheader_fill

ws_splunk['A11'] = '采购明细'
ws_splunk['A11'].font = Font(bold=True)

splunk_headers = ['序号', '产品/服务', '版本', '单位', '数量', '单价', '小计', '税率', '备注']
for col, header in enumerate(splunk_headers, 1):
    cell = ws_splunk.cell(row=12, column=col, value=header)
    set_header_style(cell)

splunk_data = [
    (1, 'Splunk Enterprise - Term License with Standard Success Plan', 'Enterprise', '套', 880, 5.89, 1891868, '13%', '第一年(2023/5/1-2024/4/30)'),
    (2, 'Splunk Enterprise - Term License with Standard Success Plan', 'Enterprise', '套', 880, 5.89, 1891868, '13%', '第二年(2024/5/1-2025/4/30)'),
    (3, 'Splunk Enterprise - Term License with Standard Success Plan', 'Enterprise', '套', 880, 5.89, 1891868, '13%', '第三年(2025/5/1-2026/4/30)'),
    ('', '不含税价', '', '', '', '', 5022658.41, '', ''),
    ('', '增值税', '', '', '', '', 652945.59, '', ''),
    ('', '含税总价', '', '', '', '', 5675604, '', ''),
]

for row, item in enumerate(splunk_data, 13):
    for col, val in enumerate(item, 1):
        cell = ws_splunk.cell(row=row, column=col, value=val)
        set_cell_style(cell)
        if row >= 15:  # 汇总行
            cell.font = Font(bold=True)

# 付款信息
ws_splunk['A20'] = '付款条件'
ws_splunk['A20'].font = Font(bold=True)
ws_splunk['A21'] = '验收合格后60日付款，需提供增值税专用发票'

# 设置列宽
for i, width in enumerate([8, 55, 12, 8, 8, 10, 12, 8, 25], 1):
    ws_splunk.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 4: Adobe 合同概要 ==========
ws_adobe = wb.create_sheet('Adobe-合同概要')

ws_adobe['A1'] = 'Adobe Experience Cloud 合同概要'
ws_adobe['A1'].font = Font(bold=True, size=14)
ws_adobe.merge_cells('A1:F1')

ws_adobe['A3'] = 'Adobe Sign Enterprise Term License and Support Agreement'
ws_adobe['A3'].font = Font(bold=True, size=12)

adobe_info = [
    ('合同类型', 'Enterprise Term License and Support Agreement'),
    ('生效日期', 'December 1, 2023'),
    ('到期日期', 'December 1, 2026'),
    ('续约条款', 'Unless otherwise indicated, any renewal term will be at the then-current list price'),
    ('支付条款', 'Net 45 from invoice date'),
    ('许可类型', 'Enterprise Named-User AL'),
    ('订购内容', 'Adobe Sign Enterprise, Adobe Sign for Microsoft, Adobe Acrobat Sign for Salesforce'),
]

for row, (label, value) in enumerate(adobe_info, 5):
    ws_adobe.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws_adobe.cell(row=row, column=1).fill = subheader_fill
    ws_adobe.cell(row=row, column=2, value=value)
    ws_adobe.merge_cells(f'B{row}:F{row}')
    for col in [1, 2]:
        set_cell_style(ws_adobe.cell(row=row, column=col))

ws_adobe['A13'] = 'Adobe Experience Manager: Cloud Service (2023v1)'
ws_adobe['A13'].font = Font(bold=True, size=12)

aem_info = [
    ('合同类型', 'Adobe Experience Manager as a Cloud Service'),
    ('客户责任', '创建和测试客户定制、提交存储处理变更内容、进行质量和安全测试'),
    ('开发顾问', '必须具备AEM开发者认证'),
    ('数据备份', 'License Term到期或终止后30天内可访问客户数据'),
    ('服务级别', 'Unified SLA (详情见 https://www.adobe.com/legal/service-commitments.html)'),
]

for row, (label, value) in enumerate(aem_info, 15):
    ws_adobe.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws_adobe.cell(row=row, column=1).fill = subheader_fill
    ws_adobe.cell(row=row, column=2, value=value)
    ws_adobe.merge_cells(f'B{row}:F{row}')
    for col in [1, 2]:
        set_cell_style(ws_adobe.cell(row=row, column=col))

# 设置列宽
ws_adobe.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F']:
    ws_adobe.column_dimensions[col].width = 25

# 保存文件
output_path = r'C:\Users\zz\Desktop\OCR Testing\合同信息汇总.xlsx'
wb.save(output_path)
print(f'Excel文件已生成: {output_path}')
